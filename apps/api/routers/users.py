import csv
import io
import json
import uuid
from datetime import datetime, timezone
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import or_
from sqlalchemy.orm import Session

from core.store import CurrentUser
from core.deps import get_current_user, require_role
from core.user_utils import user_to_current
from core.roles import ADMIN_ROLES, can_manage_role, can_manage_user
from database import get_db
from models.user import User, UserRole
from models.course import UserCourseProgress
from models.learning import UserBadge, UserCertificate

router = APIRouter(prefix="/api/v1/users", tags=["users"])


# ── Schemas ───────────────────────────────────────────────────────────────────

class UserOut(BaseModel):
    id: str
    email: str
    first_name: str
    last_name: str
    role: str
    is_verified: bool
    school: str
    bio: str
    avatar_url: Optional[str]
    linkedin: str
    github: str
    is_profile_complete: bool
    consent: dict


class UpdateProfileRequest(BaseModel):
    first_name: Optional[str] = None
    last_name: Optional[str] = None
    school: Optional[str] = None
    bio: Optional[str] = None
    avatar_url: Optional[str] = None
    linkedin: Optional[str] = None
    github: Optional[str] = None


class ConsentRequest(BaseModel):
    analytics: bool
    tracking: bool


def _to_out(cu: CurrentUser) -> UserOut:
    return UserOut(
        id=cu.id, email=cu.email,
        first_name=cu.first_name, last_name=cu.last_name,
        role=cu.role, is_verified=cu.is_verified,
        school=cu.school, bio=cu.bio,
        avatar_url=cu.avatar_url, linkedin=cu.linkedin, github=cu.github,
        is_profile_complete=cu.is_profile_complete,
        consent=cu.consent,
    )


# ── Profile ───────────────────────────────────────────────────────────────────

@router.get("/me", response_model=UserOut)
def get_me(current_user: CurrentUser = Depends(get_current_user)):
    return _to_out(current_user)


@router.put("/me", response_model=UserOut)
def update_me(
    body: UpdateProfileRequest,
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    user = db.query(User).filter(User.id == current_user.id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    for field, value in body.model_dump(exclude_none=True).items():
        setattr(user, field, value)

    db.commit()
    db.refresh(user)
    return _to_out(user_to_current(user))


# ── RGPD — Art. 15 : Droit d'accès ───────────────────────────────────────────

@router.get("/me/data")
def get_my_data(
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    user = db.query(User).filter(User.id == current_user.id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    progress_rows = db.query(UserCourseProgress).filter(
        UserCourseProgress.user_id == current_user.id
    ).all()
    badge_rows = db.query(UserBadge).filter(UserBadge.user_id == current_user.id).all()
    cert_rows = db.query(UserCertificate).filter(UserCertificate.user_id == current_user.id).all()

    return {
        "id": user.id,
        "email": user.email,
        "first_name": user.first_name or "",
        "last_name": user.last_name or "",
        "role": str(user.role.value if hasattr(user.role, "value") else user.role),
        "school": user.school or "",
        "bio": user.bio or "",
        "avatar_url": user.avatar_url,
        "linkedin": user.linkedin or "",
        "github": user.github or "",
        "created_at": user.created_at.isoformat() if user.created_at else None,
        "consent": current_user.consent,
        "_export_date": datetime.now(timezone.utc).isoformat(),
        "_platform": "Hi! Platform — Hi! PARIS",
        "learning_progress": [
            {
                "course_id": p.course_id,
                "progress_pct": p.progress_pct,
                "score": p.score,
                "completed": p.completed_at is not None,
                "completed_at": p.completed_at.isoformat() if p.completed_at else None,
            }
            for p in progress_rows
        ],
        "badges": [
            {"badge_id": b.badge_id, "awarded_at": b.awarded_at.isoformat()}
            for b in badge_rows
        ],
        "certificates": [
            {
                "id": c.id,
                "course_id": c.course_id,
                "course_title": c.course_title,
                "issued_at": c.issued_at.isoformat(),
            }
            for c in cert_rows
        ],
    }


# ── RGPD — Art. 20 : Portabilité ─────────────────────────────────────────────

@router.get("/me/export")
def export_my_data(
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    data = get_my_data(current_user=current_user, db=db)
    json_bytes = json.dumps(data, ensure_ascii=False, indent=2).encode("utf-8")
    return StreamingResponse(
        io.BytesIO(json_bytes),
        media_type="application/json",
        headers={
            "Content-Disposition": f"attachment; filename=hi-platform-export-{current_user.id[:8]}.json"
        },
    )


# ── RGPD — Art. 17 : Droit à l'effacement ────────────────────────────────────

@router.delete("/me", status_code=status.HTTP_204_NO_CONTENT)
def delete_me(
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Anonymise les données PII dans les 30 jours (ici : immédiat)."""
    user = db.query(User).filter(User.id == current_user.id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    anon_id = f"deleted_{uuid.uuid4().hex[:8]}"
    user.email = f"{anon_id}@deleted.invalid"
    user.first_name = "[supprimé]"
    user.last_name = "[supprimé]"
    user.school = None
    user.bio = None
    user.avatar_url = None
    user.linkedin = None
    user.github = None
    user.refresh_token = None
    user.is_anonymized = True
    db.commit()


# ── RGPD — Art. 21 : Consentement ────────────────────────────────────────────

@router.put("/me/consent", response_model=UserOut)
def update_consent(
    body: ConsentRequest,
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    user = db.query(User).filter(User.id == current_user.id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    user.consent_analytics = body.analytics
    user.consent_tracking = body.tracking
    user.consent_updated_at = datetime.now(timezone.utc)
    db.commit()
    db.refresh(user)
    return _to_out(user_to_current(user))


@router.get("/me/consent")
def get_consent(current_user: CurrentUser = Depends(get_current_user)):
    return current_user.consent


# ── Import massif Excel (admin) ───────────────────────────────────────────────

@router.post("/import")
async def import_users_excel(
    file: UploadFile = File(...),
    current_user: CurrentUser = Depends(require_role(*ADMIN_ROLES)),
    db: Session = Depends(get_db),
):
    """
    Import batch depuis Excel.
    Colonnes attendues: email, first_name, last_name, school, role (optionnel)
    """
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Fichier Excel (.xlsx) requis")

    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl non installé")

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active

    headers = [
        str(c.value).strip().lower() if c.value else ""
        for c in next(ws.iter_rows(min_row=1, max_row=1))
    ]
    required = {"email", "first_name", "last_name"}
    if not required.issubset(set(headers)):
        raise HTTPException(
            status_code=400,
            detail=f"Colonnes requises: email, first_name, last_name. Trouvé: {headers}",
        )

    created, skipped, errors = [], [], []

    for row in ws.iter_rows(min_row=2, values_only=True):
        row_data = {
            headers[i]: (str(v).strip() if v is not None else "")
            for i, v in enumerate(row)
        }
        email = row_data.get("email", "").lower()

        if not email or "@" not in email:
            errors.append({"email": email, "reason": "Email invalide"})
            continue

        if db.query(User).filter(User.email == email).first():
            skipped.append(email)
            continue

        role_val = row_data.get("role", "student")
        if role_val not in ("student", "teacher", "admin"):
            role_val = "student"
        # Délégation : l'importateur ne peut créer que des rôles sur lesquels il a autorité
        # (un admin ne fabrique pas d'admin ; seul un super_admin le peut).
        if not can_manage_user(current_user.role, role_val):
            role_val = "student"

        user = User(
            id=str(uuid.uuid4()),
            email=email,
            first_name=row_data.get("first_name", ""),
            last_name=row_data.get("last_name", ""),
            school=row_data.get("school", "") or None,
            role=role_val,
            hashed_password="",
            is_active=True,
            is_verified=True,
        )
        db.add(user)
        created.append(email)

    db.commit()

    return {
        "created": len(created),
        "skipped": len(skipped),
        "errors": len(errors),
        "detail": {"created": created, "skipped": skipped, "errors": errors},
    }


# ── Gestion des droits (admin / super_admin) ──────────────────────────────────
# Modèle de délégation : voir ROLES-ET-DROITS.md §6 et core/roles.py.

class AdminUserOut(BaseModel):
    id: str
    email: str
    first_name: str
    last_name: str
    role: str
    school: Optional[str]
    is_active: bool
    is_verified: bool
    created_at: Optional[str]


class UserListOut(BaseModel):
    total: int
    items: list[AdminUserOut]


class ChangeRoleRequest(BaseModel):
    role: str


class ChangeStatusRequest(BaseModel):
    is_active: bool


def _admin_out(u: User) -> AdminUserOut:
    return AdminUserOut(
        id=u.id,
        email=u.email,
        first_name=u.first_name or "",
        last_name=u.last_name or "",
        role=u.role.value if hasattr(u.role, "value") else str(u.role),
        school=u.school,
        is_active=u.is_active,
        is_verified=u.is_verified,
        created_at=u.created_at.isoformat() if u.created_at else None,
    )


def _role_value(u: User) -> str:
    return u.role.value if hasattr(u.role, "value") else str(u.role)


@router.get("", response_model=UserListOut)
def list_users(
    role: Optional[str] = None,
    school: Optional[str] = None,
    q: Optional[str] = None,
    limit: int = 50,
    offset: int = 0,
    current_user: CurrentUser = Depends(require_role(*ADMIN_ROLES)),
    db: Session = Depends(get_db),
):
    """Liste paginée des utilisateurs (admin+). Filtres : rôle, école, recherche texte."""
    query = db.query(User)
    if role:
        query = query.filter(User.role == role)
    if school:
        query = query.filter(User.school == school)
    if q:
        like = f"%{q.strip()}%"
        query = query.filter(
            or_(
                User.email.ilike(like),
                User.first_name.ilike(like),
                User.last_name.ilike(like),
            )
        )

    total = query.count()
    limit = max(1, min(limit, 200))
    rows = (
        query.order_by(User.created_at.desc())
        .offset(max(0, offset))
        .limit(limit)
        .all()
    )
    return UserListOut(total=total, items=[_admin_out(u) for u in rows])


@router.patch("/{user_id}/role", response_model=AdminUserOut)
def change_user_role(
    user_id: str,
    body: ChangeRoleRequest,
    current_user: CurrentUser = Depends(require_role(*ADMIN_ROLES)),
    db: Session = Depends(get_db),
):
    """
    Change le rôle d'un utilisateur (délégation §6).
    Un admin gère jusqu'à `teacher` ; seul un super_admin fabrique/retire des admin.
    Le dernier super_admin ne peut pas être rétrogradé.
    """
    valid_roles = {r.value for r in UserRole}
    new_role = body.role
    if new_role not in valid_roles:
        raise HTTPException(status_code=400, detail=f"Rôle invalide : {new_role}")

    target = db.query(User).filter(User.id == user_id).first()
    if not target:
        raise HTTPException(status_code=404, detail="Utilisateur introuvable")

    target_role = _role_value(target)
    if not can_manage_role(current_user.role, target_role, new_role):
        raise HTTPException(
            status_code=403,
            detail="Vous n'avez pas l'autorité pour attribuer ce rôle à cet utilisateur.",
        )

    # Garde-fou : ne pas rétrograder le dernier super_admin (plateforme orpheline).
    if target_role == UserRole.super_admin.value and new_role != UserRole.super_admin.value:
        count = db.query(User).filter(User.role == UserRole.super_admin).count()
        if count <= 1:
            raise HTTPException(
                status_code=400,
                detail="Impossible de rétrograder le dernier super_admin.",
            )

    target.role = UserRole(new_role)
    db.commit()
    db.refresh(target)
    return _admin_out(target)


@router.patch("/{user_id}/status", response_model=AdminUserOut)
def change_user_status(
    user_id: str,
    body: ChangeStatusRequest,
    current_user: CurrentUser = Depends(require_role(*ADMIN_ROLES)),
    db: Session = Depends(get_db),
):
    """Suspend (is_active=False) ou réactive un compte (admin+)."""
    target = db.query(User).filter(User.id == user_id).first()
    if not target:
        raise HTTPException(status_code=404, detail="Utilisateur introuvable")

    if target.id == current_user.id:
        raise HTTPException(status_code=400, detail="Vous ne pouvez pas suspendre votre propre compte.")

    target_role = _role_value(target)
    if not can_manage_user(current_user.role, target_role):
        raise HTTPException(status_code=403, detail="Vous n'avez pas l'autorité sur ce compte.")

    # Garde-fou : ne pas suspendre le dernier super_admin actif.
    if not body.is_active and target_role == UserRole.super_admin.value:
        active = db.query(User).filter(
            User.role == UserRole.super_admin, User.is_active == True  # noqa: E712
        ).count()
        if active <= 1:
            raise HTTPException(status_code=400, detail="Impossible de suspendre le dernier super_admin actif.")

    target.is_active = body.is_active
    db.commit()
    db.refresh(target)
    return _admin_out(target)
